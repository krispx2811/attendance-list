"""Tests for the storage and export layers (no GUI required)."""

from __future__ import annotations

import os
import sys
import tempfile
import unittest
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))


class CoreTests(unittest.TestCase):
    def setUp(self):
        self._tmp = tempfile.TemporaryDirectory()
        os.environ["ATTENDANCE_DATA_DIR"] = self._tmp.name

        from app import db

        db.close()
        self.db = db
        self.db.connect()

    def tearDown(self):
        self.db.close()
        os.environ.pop("ATTENDANCE_DATA_DIR", None)
        self._tmp.cleanup()

    # -- people ----------------------------------------------------------
    def test_add_person_is_idempotent_and_case_insensitive(self):
        first = self.db.add_person("Amina Hassan")
        second = self.db.add_person("amina hassan")
        self.assertEqual(first, second)
        self.assertEqual(len(self.db.list_people()), 1)

    def test_readding_a_removed_person_reactivates_them(self):
        pid = self.db.add_person("Sam")
        self.db.set_person_active(pid, False)
        self.assertEqual(len(self.db.list_people()), 0)

        self.assertEqual(self.db.add_person("Sam"), pid)
        self.assertEqual(len(self.db.list_people()), 1)

    def test_rename_rejects_a_duplicate_name(self):
        self.db.add_person("Alice")
        bob = self.db.add_person("Bob")
        with self.assertRaises(ValueError):
            self.db.rename_person(bob, "alice")

    def test_empty_name_is_rejected(self):
        with self.assertRaises(ValueError):
            self.db.add_person("   ")

    # -- marking ---------------------------------------------------------
    def test_remarking_updates_instead_of_duplicating(self):
        pid = self.db.add_person("Dana")
        day = "2026-08-03"

        self.db.mark(pid, day, self.db.STATUS_ABSENT, "Sick")
        self.db.mark(pid, day, self.db.STATUS_LATE, "Traffic")

        rows = self.db.search()
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["status"], self.db.STATUS_LATE)
        self.assertEqual(rows[0]["reason"], "Traffic")

    def test_reason_is_cleared_when_marked_present(self):
        pid = self.db.add_person("Eli")
        day = "2026-08-03"

        self.db.mark(pid, day, self.db.STATUS_ABSENT, "Sick")
        self.db.mark(pid, day, self.db.STATUS_PRESENT, "Sick")

        self.assertEqual(self.db.search()[0]["reason"], "")

    def test_unknown_status_is_rejected(self):
        pid = self.db.add_person("Fay")
        with self.assertRaises(ValueError):
            self.db.mark(pid, "2026-08-03", "Maybe", "")

    def test_unmark_removes_the_record(self):
        pid = self.db.add_person("Gus")
        self.db.mark(pid, "2026-08-03", self.db.STATUS_PRESENT)
        self.db.unmark(pid, "2026-08-03")
        self.assertEqual(self.db.search(), [])

    # -- day view --------------------------------------------------------
    def test_get_day_lists_roster_plus_that_days_walkins_only(self):
        roster = self.db.add_person("Roster Person")
        walkin = self.db.add_person("Walk In", kind=self.db.KIND_WALKIN)
        self.db.mark(walkin, "2026-08-03", self.db.STATUS_PRESENT)

        marked_day = {r["name"] for r in self.db.get_day("2026-08-03")}
        other_day = {r["name"] for r in self.db.get_day("2026-08-04")}

        self.assertEqual(marked_day, {"Roster Person", "Walk In"})
        self.assertEqual(other_day, {"Roster Person"})
        self.assertTrue(roster)

    def test_removed_person_still_shows_on_days_they_were_marked(self):
        pid = self.db.add_person("Past Employee")
        self.db.mark(pid, "2026-07-01", self.db.STATUS_ABSENT, "Holiday")
        self.db.set_person_active(pid, False)

        names = {r["name"] for r in self.db.get_day("2026-07-01")}
        self.assertIn("Past Employee", names)

    def test_day_summary_counts_unrecorded_people(self):
        a = self.db.add_person("A")
        self.db.add_person("B")
        self.db.mark(a, "2026-08-03", self.db.STATUS_ABSENT, "Sick")

        summary = self.db.day_summary("2026-08-03")
        self.assertEqual(summary[self.db.STATUS_ABSENT], 1)
        self.assertEqual(summary["Unrecorded"], 1)
        self.assertEqual(summary["Total"], 2)

    # -- search / stats --------------------------------------------------
    def test_search_filters_by_name_range_and_status(self):
        a = self.db.add_person("Hana")
        b = self.db.add_person("Omar")
        self.db.mark(a, "2026-08-01", self.db.STATUS_ABSENT, "Sick")
        self.db.mark(a, "2026-08-05", self.db.STATUS_PRESENT)
        self.db.mark(b, "2026-08-05", self.db.STATUS_ABSENT, "Sick")

        self.assertEqual(len(self.db.search(name_query="han")), 2)
        self.assertEqual(len(self.db.search(start="2026-08-02")), 2)
        self.assertEqual(len(self.db.search(end="2026-08-01")), 1)
        self.assertEqual(len(self.db.search(status=self.db.STATUS_ABSENT)), 2)

    def test_per_person_stats_include_people_with_no_records_in_range(self):
        a = self.db.add_person("Ines")
        self.db.add_person("Nobody")
        self.db.mark(a, "2026-08-01", self.db.STATUS_PRESENT)
        self.db.mark(a, "2026-08-02", self.db.STATUS_ABSENT, "Sick")
        self.db.mark(a, "2026-08-03", self.db.STATUS_LATE, "Bus")

        stats = {s["name"]: s for s in self.db.per_person_stats()}
        self.assertEqual(stats["Ines"]["recorded"], 3)
        self.assertEqual(stats["Ines"]["absent"], 1)
        self.assertEqual(stats["Nobody"]["recorded"], 0)

    def test_stats_respect_the_date_range(self):
        a = self.db.add_person("Jo")
        self.db.mark(a, "2026-07-01", self.db.STATUS_PRESENT)
        self.db.mark(a, "2026-08-01", self.db.STATUS_ABSENT, "Sick")

        stats = {s["name"]: s for s in self.db.per_person_stats(start="2026-08-01")}
        self.assertEqual(stats["Jo"]["recorded"], 1)
        self.assertEqual(stats["Jo"]["absent"], 1)

    def test_known_reasons_are_ordered_by_frequency(self):
        a = self.db.add_person("K")
        b = self.db.add_person("L")
        self.db.mark(a, "2026-08-01", self.db.STATUS_ABSENT, "Sick")
        self.db.mark(b, "2026-08-01", self.db.STATUS_ABSENT, "Sick")
        self.db.mark(a, "2026-08-02", self.db.STATUS_ABSENT, "Holiday")

        self.assertEqual(self.db.known_reasons()[0], "Sick")

    # -- default employees -----------------------------------------------
    def test_a_new_database_is_seeded_with_the_default_employees(self):
        added = self.db.seed_default_people()
        self.assertEqual(added, len(self.db.DEFAULT_EMPLOYEES))

        names = [p["name"] for p in self.db.list_people()]
        for expected in self.db.DEFAULT_EMPLOYEES:
            self.assertIn(expected, names)

    def test_seeding_is_skipped_once_anyone_exists(self):
        self.db.add_person("Only Me")
        self.assertEqual(self.db.seed_default_people(), 0)
        self.assertEqual([p["name"] for p in self.db.list_people()], ["Only Me"])

    def test_seeding_does_not_resurrect_a_removed_default(self):
        self.db.seed_default_people()
        target = self.db.list_people()[0]
        self.db.set_person_active(target["id"], False)

        self.db.seed_default_people()  # e.g. after an app restart
        active = [p["name"] for p in self.db.list_people()]
        self.assertNotIn(target["name"], active)

    def test_default_names_are_stored_exactly_as_written(self):
        self.db.seed_default_people()
        names = [p["name"] for p in self.db.list_people()]
        self.assertIn("Sara ahmed al balushi", names)
        self.assertIn("Sara al balushi", names)

    # -- deletion / backup -----------------------------------------------
    def test_deleting_a_person_cascades_to_their_records(self):
        pid = self.db.add_person("Mia")
        self.db.mark(pid, "2026-08-03", self.db.STATUS_PRESENT)
        self.db.delete_person(pid)
        self.assertEqual(self.db.search(), [])

    def test_backup_creates_a_readable_copy(self):
        pid = self.db.add_person("Noor")
        self.db.mark(pid, "2026-08-03", self.db.STATUS_ABSENT, "Sick")

        target = self.db.backup_now()
        self.assertIsNotNone(target)
        self.assertTrue(Path(target).exists())

        import sqlite3

        conn = sqlite3.connect(str(target))
        count = conn.execute("SELECT COUNT(*) FROM attendance").fetchone()[0]
        conn.close()
        self.assertEqual(count, 1)

    # -- exports ---------------------------------------------------------
    def test_csv_and_xlsx_exports_contain_the_records(self):
        from app import exporter

        pid = self.db.add_person("Priya")
        self.db.mark(pid, "2026-08-03", self.db.STATUS_ABSENT, "Doctor's appointment")
        rows = self.db.search()

        csv_path = Path(self._tmp.name) / "out.csv"
        exporter.export_csv(rows, csv_path)
        text = csv_path.read_text(encoding="utf-8-sig")
        self.assertIn("Priya", text)
        self.assertIn("Doctor's appointment", text)

        xlsx_path = Path(self._tmp.name) / "out.xlsx"
        exporter.export_xlsx(rows, xlsx_path)
        self.assertTrue(xlsx_path.exists())

        from openpyxl import load_workbook

        ws = load_workbook(xlsx_path).active
        values = [c.value for c in ws[2]]
        self.assertIn("Priya", values)
        self.assertIn("Absent", values)

    def test_summary_workbook_has_three_sheets(self):
        from app import exporter
        from openpyxl import load_workbook

        pid = self.db.add_person("Rami")
        self.db.mark(pid, "2026-08-03", self.db.STATUS_ABSENT, "Sick")

        target = Path(self._tmp.name) / "report.xlsx"
        exporter.export_summary_xlsx(
            self.db.search(), self.db.per_person_stats(), self.db.reason_stats(), target
        )

        wb = load_workbook(target)
        self.assertEqual(wb.sheetnames, ["Records", "Summary", "Reasons"])


class PathTests(unittest.TestCase):
    """The data folder sits beside the app unless that folder is read-only."""

    def setUp(self):
        from app import paths

        self.paths = paths
        os.environ.pop("ATTENDANCE_DATA_DIR", None)
        paths._resolved_data_dir = None

    def tearDown(self):
        self.paths._resolved_data_dir = None
        os.environ.pop("ATTENDANCE_DATA_DIR", None)

    def test_data_lives_beside_the_app_by_default(self):
        self.assertEqual(self.paths.data_dir(), self.paths.app_dir() / "data")
        self.assertFalse(self.paths.using_fallback_location())

    def test_env_override_wins(self):
        with tempfile.TemporaryDirectory() as tmp:
            os.environ["ATTENDANCE_DATA_DIR"] = tmp
            self.assertEqual(self.paths.data_dir(), Path(tmp))

    def test_falls_back_when_the_app_folder_is_not_writable(self):
        original = self.paths._is_writable
        self.paths._is_writable = lambda _p: False
        try:
            self.paths._resolved_data_dir = None
            self.assertEqual(self.paths.data_dir(), self.paths.user_data_dir())
            self.assertTrue(self.paths.using_fallback_location())
        finally:
            self.paths._is_writable = original

    def test_legacy_database_is_adopted_and_the_original_kept(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            new_dir, legacy_dir = root / "new", root / "legacy"
            new_dir.mkdir()
            legacy_dir.mkdir()

            # A legacy database with one identifiable row.
            import sqlite3

            legacy_db = legacy_dir / "attendance.db"
            conn = sqlite3.connect(str(legacy_db))
            conn.execute("CREATE TABLE marker (note TEXT)")
            conn.execute("INSERT INTO marker VALUES ('carried over')")
            conn.commit()
            conn.close()

            original_dir, original_user = self.paths.data_dir, self.paths.user_data_dir
            self.paths.data_dir = lambda: new_dir
            self.paths.user_data_dir = lambda: legacy_dir
            try:
                adopted = self.paths.adopt_legacy_database()
                self.assertIsNotNone(adopted)

                conn = sqlite3.connect(str(new_dir / "attendance.db"))
                note = conn.execute("SELECT note FROM marker").fetchone()[0]
                conn.close()
                self.assertEqual(note, "carried over")

                # Original preserved under a new name, not deleted.
                self.assertFalse(legacy_db.exists())
                self.assertTrue((legacy_dir / "attendance.db.migrated").exists())

                # Running again must not overwrite the live database.
                self.assertIsNone(self.paths.adopt_legacy_database())
            finally:
                self.paths.data_dir = original_dir
                self.paths.user_data_dir = original_user

    def test_adoption_carries_rows_still_sitting_in_the_wal(self):
        """A plain file copy would lose these; the backup API must not."""
        import sqlite3

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            new_dir, legacy_dir = root / "new", root / "legacy"
            new_dir.mkdir()
            legacy_dir.mkdir()

            legacy_db = legacy_dir / "attendance.db"
            conn = sqlite3.connect(str(legacy_db))
            conn.execute("PRAGMA journal_mode = WAL")
            conn.execute("CREATE TABLE marker (note TEXT)")
            conn.execute("INSERT INTO marker VALUES ('written in wal mode')")
            conn.commit()
            # Deliberately leave the connection open so the -wal is not
            # checkpointed into the main file, mimicking an app still running.

            original_dir, original_user = self.paths.data_dir, self.paths.user_data_dir
            self.paths.data_dir = lambda: new_dir
            self.paths.user_data_dir = lambda: legacy_dir
            try:
                self.assertTrue(legacy_db.with_name("attendance.db-wal").exists())
                self.assertIsNotNone(self.paths.adopt_legacy_database())

                check = sqlite3.connect(str(new_dir / "attendance.db"))
                note = check.execute("SELECT note FROM marker").fetchone()[0]
                check.close()
                self.assertEqual(note, "written in wal mode")
            finally:
                self.paths.data_dir = original_dir
                self.paths.user_data_dir = original_user


class VersionTests(unittest.TestCase):
    def test_version_comparison_handles_tags_and_suffixes(self):
        from app.version import version_tuple

        self.assertEqual(version_tuple("v1.2.3"), (1, 2, 3))
        self.assertEqual(version_tuple("1.2"), (1, 2, 0))
        self.assertEqual(version_tuple("1.2.3-beta"), (1, 2, 3))
        self.assertGreater(version_tuple("v1.10.0"), version_tuple("v1.9.0"))
        self.assertGreater(version_tuple("2.0.0"), version_tuple("1.99.99"))


if __name__ == "__main__":
    unittest.main(verbosity=2)
