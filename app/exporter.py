"""CSV and Excel export.

Both formats share one column layout so a spreadsheet built from either is
interchangeable.
"""

from __future__ import annotations

import csv
import sqlite3
from pathlib import Path
from typing import Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from . import db

HEADERS = ["Date", "Name", "Status", "Reason for not coming", "Recorded at"]

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(color="FFFFFF", bold=True)

_STATUS_FILL = {
    db.STATUS_PRESENT: PatternFill("solid", fgColor="E2F0D9"),
    db.STATUS_LATE: PatternFill("solid", fgColor="FFF2CC"),
    db.STATUS_ABSENT: PatternFill("solid", fgColor="FBE5E5"),
}


def _rows_to_records(rows: Sequence[sqlite3.Row]) -> list[list[str]]:
    return [
        [
            r["date"],
            r["name"],
            r["status"],
            r["reason"] or "",
            r["recorded_at"],
        ]
        for r in rows
    ]


def export_csv(rows: Sequence[sqlite3.Row], destination: Path) -> Path:
    destination = Path(destination)
    # utf-8-sig so Excel on Windows opens accented names correctly.
    with destination.open("w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.writer(fh)
        writer.writerow(HEADERS)
        writer.writerows(_rows_to_records(rows))
    return destination


def export_xlsx(
    rows: Sequence[sqlite3.Row], destination: Path, sheet_title: str = "Attendance"
) -> Path:
    destination = Path(destination)
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31] or "Attendance"

    ws.append(HEADERS)
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for record in _rows_to_records(rows):
        ws.append(record)

    # Tint each row by status so absences stand out when scanning.
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        fill = _STATUS_FILL.get(row[2].value)
        if fill:
            for cell in row:
                cell.fill = fill

    _autosize(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{ws.max_row}"

    wb.save(destination)
    return destination


def export_summary_xlsx(
    rows: Sequence[sqlite3.Row],
    stats: Sequence[sqlite3.Row],
    reasons: Sequence[sqlite3.Row],
    destination: Path,
) -> Path:
    """Workbook with the raw records plus a per-person summary and reasons."""
    destination = Path(destination)
    wb = Workbook()

    ws = wb.active
    ws.title = "Records"
    ws.append(HEADERS)
    for record in _rows_to_records(rows):
        ws.append(record)
    _style_header(ws)
    _autosize(ws)
    ws.freeze_panes = "A2"

    summary = wb.create_sheet("Summary")
    summary.append(["Name", "Recorded days", "Present", "Late", "Absent", "Attendance rate"])
    for s in stats:
        recorded = s["recorded"] or 0
        attended = (s["present"] or 0) + (s["late"] or 0)
        rate = (attended / recorded) if recorded else 0
        summary.append(
            [
                s["name"],
                recorded,
                s["present"] or 0,
                s["late"] or 0,
                s["absent"] or 0,
                rate,
            ]
        )
    for row in summary.iter_rows(min_row=2, max_row=summary.max_row, min_col=6, max_col=6):
        for cell in row:
            cell.number_format = "0%"
    _style_header(summary)
    _autosize(summary)
    summary.freeze_panes = "A2"

    reason_sheet = wb.create_sheet("Reasons")
    reason_sheet.append(["Reason", "Times given"])
    for r in reasons:
        reason_sheet.append([r["reason"], r["n"]])
    _style_header(reason_sheet)
    _autosize(reason_sheet)

    wb.save(destination)
    return destination


def _style_header(ws) -> None:
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _autosize(ws, minimum: int = 10, maximum: int = 45) -> None:
    for column in ws.columns:
        longest = max((len(str(c.value)) for c in column if c.value is not None), default=0)
        letter = get_column_letter(column[0].column)
        ws.column_dimensions[letter].width = max(minimum, min(maximum, longest + 3))
